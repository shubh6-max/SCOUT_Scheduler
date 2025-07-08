import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import os
import pytz
import send_mails  # must contain send_mails.main()

# --- Config ---
EXCEL_FILE = "https://themathcompany.sharepoint.com/sites/scout/_layouts/15/guestaccess.aspx?share=Easz2VQSTFtIvnldo0onHUIBssWNAxO_HPLhwRT8L2sS-w&e=01roDe"
SHEET_NAME = "Sheet1"
SENT_LOG_FILE = "last_email_sent.txt"
IST = pytz.timezone("Asia/Kolkata")

# --- Scheduler logic ---
def get_ist_time():
    now = datetime.now(IST)
    return now.strftime("%A"), now.strftime("%H:%M"), now.strftime("%Y-%m-%d")

def already_sent_today(today):
    if os.path.exists(SENT_LOG_FILE):
        with open(SENT_LOG_FILE, "r") as f:
            return f.read().strip() == today
    return False

def mark_sent_today(today):
    with open(SENT_LOG_FILE, "w") as f:
        f.write(today)

# --- Run email scheduler logic ---
day, time, today_str = get_ist_time()
if day == "Tuesday" and time == "23:40":
    if not already_sent_today(today_str):
        send_mails.main()
        mark_sent_today(today_str)

# --- Streamlit Config ---
st.set_page_config(page_title="Warm Outreach Form", layout="centered")

# --- Load Excel ---
@st.cache_data
def load_data():
    return pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)

df = load_data()

# --- Get stakeholder email from query param ---
query_params = st.query_params
stakeholder_email = query_params.get("email", None)

st.title("🌟 Warm Outreach Relationship Survey")

if stakeholder_email is None:
    st.warning("Please access this form using your personalized email link (e.g., ?email=your@email.com)")
    st.stop()

# --- Filter for this stakeholder & not done leads ---
filtered_df = df[
    df["Leadership contact email"].str.contains(stakeholder_email, case=False, na=False) &
    (df["Status"].str.lower() == "not done")
]

if filtered_df.empty:
    st.info("✅ You have no pending leads. All caught up!")
    st.stop()

st.markdown(f"**Welcome, {stakeholder_email}**")
st.markdown("Please rate your relationship strength with the following leads:")

# --- Form starts ---
responses = []
with st.form("relationship_form"):
    for idx, row in filtered_df.iterrows():
        lead_name = row["Target Lead Name"]
        linkedin_url = row["Target Lead Linkedin URL"]

        st.markdown(f"### {lead_name} — [LinkedIn Profile]({linkedin_url})")

        score = st.radio(
            f"What is your relationship strength with {lead_name}?",
            options=[
                "1 - Don’t Know",
                "2 - Met Once",
                "3 - Professional Acquaintance",
                "4 - Regular Contact",
                "5 - Close Relationship"
            ],
            key=f"{lead_name}_score"
        )

        comment = st.text_input(f"Comments for {lead_name} (optional)", key=f"{lead_name}_comment")

        responses.append({
            "row_index": idx,
            "lead": lead_name,
            "score": score,
            "comment": comment
        })

    submitted = st.form_submit_button("Submit Responses")

# --- Save logic ---
if submitted:
    score_col = f"{stakeholder_email}_Score"
    comment_col = f"{stakeholder_email}_Comment"

    # Load workbook
    book = load_workbook(EXCEL_FILE)
    sheet = book[SHEET_NAME]
    header = [cell.value for cell in sheet[1]]

    # Add columns if missing
    if score_col not in header:
        sheet.cell(row=1, column=len(header)+1, value=score_col)
        header.append(score_col)
    if comment_col not in header:
        sheet.cell(row=1, column=len(header)+1, value=comment_col)
        header.append(comment_col)

    score_col_idx = header.index(score_col) + 1
    comment_col_idx = header.index(comment_col) + 1
    status_col_idx = header.index("Status") + 1

    for resp in responses:
        row_excel = resp["row_index"] + 2
        sheet.cell(row=row_excel, column=score_col_idx, value=resp["score"])
        sheet.cell(row=row_excel, column=comment_col_idx, value=resp["comment"])
        sheet.cell(row=row_excel, column=status_col_idx, value="Done")

    book.save(EXCEL_FILE)
    st.success("✅ Your responses have been recorded. Thank you!")
